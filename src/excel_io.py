import shutil
from pathlib import Path

from openpyxl import load_workbook

from .config import (
    MAX_PHRASE_LEN,
    MIN_PHRASE_LEN,
    OPERATION_COLUMN_HEADERS,
    PHRASES_CACHE_VERSION,
    ROOT,
)
from .detect import is_abstract_phrase, is_metadata_line

PHRASES_CACHE = ROOT / "phrases_cache.json"


def iter_text_cells(workbook_path):
    wb = load_workbook(workbook_path, data_only=False)
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        for row in ws.iter_rows():
            for cell in row:
                if cell.value is None:
                    continue
                if not isinstance(cell.value, str):
                    continue
                text = cell.value.strip()
                if MIN_PHRASE_LEN <= len(text) <= MAX_PHRASE_LEN:
                    yield sheet_name, cell.coordinate, text


def _sheet_pair(sheet_name: str) -> str | None:
    if sheet_name.startswith("Лист"):
        suffix = sheet_name[4:]
        return f"Лист{suffix} ↔ {suffix}"
    if sheet_name.isdigit():
        return f"Лист{sheet_name} ↔ {sheet_name}"
    return None


def _find_operation_columns(ws) -> list[int]:
    cols: list[int] = []
    for row in ws.iter_rows(max_row=30):
        for cell in row:
            if not isinstance(cell.value, str):
                continue
            header = cell.value.strip().lower()
            if header in OPERATION_COLUMN_HEADERS:
                cols.append(cell.column - 1)
    return sorted(set(cols))


def _collect_operation_column_phrases(workbook_path) -> dict[str, set[str]]:
    wb = load_workbook(workbook_path, read_only=True, data_only=True)
    phrase_sheets: dict[str, set[str]] = {}

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        op_cols = _find_operation_columns(ws)
        if not op_cols:
            continue

        for row in ws.iter_rows(values_only=True):
            for col_idx in op_cols:
                if col_idx >= len(row):
                    continue
                val = row[col_idx]
                if not isinstance(val, str):
                    continue
                text = val.strip()
                if not text or is_metadata_line(text):
                    continue
                if not (MIN_PHRASE_LEN <= len(text) <= MAX_PHRASE_LEN):
                    continue
                phrase_sheets.setdefault(text, set()).add(sheet_name)

    return phrase_sheets


def collect_phrase_index(workbook_path) -> dict[str, dict]:
    index: dict[str, dict] = {}

    for sheet_name, _, text in iter_text_cells(workbook_path):
        if not is_abstract_phrase(text):
            continue
        entry = index.setdefault(
            text,
            {"sheets": set(), "count": 0, "in_summary": False, "in_workday": False},
        )
        entry["sheets"].add(sheet_name)
        entry["count"] += 1
        if sheet_name.startswith("Лист"):
            entry["in_summary"] = True
        elif sheet_name.isdigit():
            entry["in_workday"] = True

    for text, sheets in _collect_operation_column_phrases(workbook_path).items():
        entry = index.setdefault(
            text,
            {"sheets": set(), "count": 0, "in_summary": False, "in_workday": False},
        )
        entry["sheets"].update(sheets)
        if any(s.startswith("Лист") for s in sheets):
            entry["in_summary"] = True
        if any(s.isdigit() for s in sheets):
            entry["in_workday"] = True

    for text, entry in index.items():
        entry["sheets"] = sorted(entry["sheets"])
        pairs = sorted({p for s in entry["sheets"] if (p := _sheet_pair(s))})
        entry["pairs"] = pairs
        entry["is_duplicate_across_pair"] = entry["in_summary"] and entry["in_workday"]

    return index


def _scan_phrases(workbook_path) -> list[str]:
    return sorted(collect_phrase_index(workbook_path).keys(), key=str.lower)


def collect_unique_phrases(
    workbook_path,
    use_cache: bool = True,
    cache_path: Path | None = None,
) -> list[str]:
    cache_file = cache_path or PHRASES_CACHE
    if use_cache and cache_file.exists():
        import json

        data = json.loads(cache_file.read_text(encoding="utf-8"))
        if (
            data.get("source") == str(workbook_path)
            and data.get("version") == PHRASES_CACHE_VERSION
        ):
            return data["phrases"]

    phrases = _scan_phrases(workbook_path)

    import json

    cache_file.parent.mkdir(parents=True, exist_ok=True)
    cache_file.write_text(
        json.dumps(
            {
                "source": str(workbook_path),
                "version": PHRASES_CACHE_VERSION,
                "phrases": phrases,
            },
            ensure_ascii=False,
            indent=2,
        ),
        encoding="utf-8",
    )
    return phrases


def apply_replacements(workbook_path, output_path, replacements: dict[str, str]) -> int:
    if not replacements:
        return 0

    wb = load_workbook(workbook_path)
    count = 0
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                if isinstance(cell.value, str) and cell.value in replacements:
                    new_value = replacements[cell.value]
                    if new_value != cell.value:
                        cell.value = new_value
                        count += 1

    wb.save(output_path)
    return count


def copy_workbook(source, destination):
    try:
        shutil.copy2(source, destination)
    except PermissionError:
        alt = destination.with_name(destination.stem + "_pending" + destination.suffix)
        shutil.copy2(source, alt)
        raise PermissionError(
            f"Cannot write {destination.name} — file is open in Excel. "
            f"Close it and rerun, or use {alt.name}"
        ) from None
