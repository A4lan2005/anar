"""Auto-process Книга2.xlsx with Gemini and save as auto.xlsx."""

import argparse
import json
import sys
import time
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent))

from src.config import AUTO_XLSX, CANDIDATES_JSON, MANUAL_XLSX, SOURCE_XLSX
from src.excel_io import apply_replacements, collect_unique_phrases, copy_workbook
from src.gemini_rewrite import suggest_replacements_batch


def extract_context(workbook_path) -> str:
    from openpyxl import load_workbook

    wb = load_workbook(workbook_path, read_only=True, data_only=True)
    for name in wb.sheetnames[:3]:
        ws = wb[name]
        for row in ws.iter_rows(max_row=10, values_only=True):
            for val in row:
                if isinstance(val, str) and "Наименование работы" in val:
                    return val[:500]
    return ""


def build_replacements(phrases: list[str], cache: dict) -> dict[str, str]:
    replacements: dict[str, str] = {}
    for phrase in phrases:
        if phrase in cache and cache[phrase].get("candidates"):
            rec = cache[phrase]["recommended"]
            replacements[phrase] = cache[phrase]["candidates"][rec]
    return replacements


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--fresh", action="store_true", help="Recreate manual.xlsx and auto.xlsx from source")
    parser.add_argument("--apply-only", action="store_true", help="Apply cached candidates to auto.xlsx only")
    args = parser.parse_args()

    if not SOURCE_XLSX.exists():
        print(f"Source not found: {SOURCE_XLSX}")
        sys.exit(1)

    if args.fresh or not MANUAL_XLSX.exists() or not AUTO_XLSX.exists():
        print("Creating copies: manual.xlsx, auto.xlsx ...", flush=True)
        copy_workbook(SOURCE_XLSX, MANUAL_XLSX)
        copy_workbook(SOURCE_XLSX, AUTO_XLSX)
    else:
        print("Using existing manual.xlsx and auto.xlsx", flush=True)

    phrases = collect_unique_phrases(SOURCE_XLSX)
    print(f"Found {len(phrases)} unique abstract phrases", flush=True)

    cache: dict = {}
    if CANDIDATES_JSON.exists():
        cache = json.loads(CANDIDATES_JSON.read_text(encoding="utf-8"))

    if args.apply_only:
        replacements = build_replacements(phrases, cache)
        print(f"Applying {len(replacements)} cached replacements to auto.xlsx ...", flush=True)
        copy_workbook(SOURCE_XLSX, AUTO_XLSX)
        count = apply_replacements(AUTO_XLSX, AUTO_XLSX, replacements)
        print(f"Done. Updated {count} cells in {AUTO_XLSX}", flush=True)
        return

    context = extract_context(SOURCE_XLSX)
    batch_size = 15
    pending = [p for p in phrases if p not in cache]

    for i in range(0, len(pending), batch_size):
        batch = pending[i : i + batch_size]
        batch_num = i // batch_size + 1
        total_batches = (len(pending) + batch_size - 1) // batch_size
        print(
            f"Gemini batch {batch_num}/{total_batches}: {len(batch)} phrases "
            f"(cached {len(cache)}/{len(phrases)})",
            flush=True,
        )
        try:
            results = suggest_replacements_batch(batch, context)
            cache.update(results)
            CANDIDATES_JSON.write_text(json.dumps(cache, ensure_ascii=False, indent=2), encoding="utf-8")
        except Exception as exc:
            print(f"  Batch failed: {exc}, retrying one-by-one ...", flush=True)
            from src.gemini_rewrite import suggest_replacements

            for phrase in batch:
                try:
                    cache[phrase] = suggest_replacements(phrase, context)
                except Exception as e2:
                    print(f"  Skip {phrase[:50]}...: {e2}", flush=True)
                time.sleep(0.5)
            CANDIDATES_JSON.write_text(json.dumps(cache, ensure_ascii=False, indent=2), encoding="utf-8")
        time.sleep(0.5)

    replacements = build_replacements(phrases, cache)
    print(f"Applying {len(replacements)} replacements to auto.xlsx ...", flush=True)
    copy_workbook(SOURCE_XLSX, AUTO_XLSX)
    count = apply_replacements(AUTO_XLSX, AUTO_XLSX, replacements)
    print(f"Done. Updated {count} cells in {AUTO_XLSX}", flush=True)


if __name__ == "__main__":
    main()
